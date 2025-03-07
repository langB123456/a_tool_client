import os, time, datetime, random, unicodedata, re, PyPDF2, psutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


# 杀死进程
def kill_processes_by_names(process_names):
    for proc in psutil.process_iter(['name']):
        try:
            if proc.info['name'] in process_names:
                proc.kill()
        except psutil.NoSuchProcess:
            print(f"进程 {proc.info['name']} 不存在.")
            continue
        except psutil.AccessDenied:
            print(f"无法终止进程 {proc.info['name']}.")


# 获取pdf数据
def get_pdf_data(pdf_path):
    # 打开PDF文件
    pdf_file = open(pdf_path, 'rb')
    # 创建一个PDF对象
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    # 获取PDF文件中的页面数量
    num_pages = len(pdf_reader.pages)
    # 创建一个空字符串，用于存储提取的文本
    text = ""
    # 循环遍历每一页并提取文本
    for page_num in range(num_pages):
        page = pdf_reader.pages[page_num]
        text += page.extract_text()
    # 关闭PDF文件
    pdf_file.close()
    # 打印提取的文本
    pdf_data = text.split("\n")
    return pdf_data


# 获取件数 pdf_num
def get_num(data):
    try:
        for i in range(len(data)):
            if re.search(r'Total[  ]packages', data[i]):
                if len(data[i]) > 30:
                    origin_data = data[i]
                    if '\xa0' in origin_data:
                        print(origin_data)
                        if is_number(data[i].split("\xa0")[-1]):
                            return data[i].split("\xa0")[-1]
                        else:
                            return data[i].split("\xa0")[-1][6:]
                    if ' ' in origin_data:
                        if is_number(data[i].split(" ")[-1]):
                            return data[i].split(" ")[-1]
                        else:
                            return data[i].split(" ")[-1][6:]
                else:
                    return data[i + 1]
    except Exception as e:
        print(e)


# 获取税号 pdf_vat
def get_Tax_ID_old(data):
    pattern = r'GB\d{12}'
    Tax_ID_list = []
    for i in range(len(data)):
        if re.search(pattern, data[i]):
            result = re.findall(pattern, data[i])
            Tax_ID_list = Tax_ID_list + result
    if len(Tax_ID_list) == 1:
        return Tax_ID_list[0].replace(" ", "")
    else:
        return Tax_ID_list[1].replace(" ", "")


# 获取税号 pdf_vat
def get_Tax_ID(data):
    pattern = r'GB\d{12}'
    Tax_ID_list = []
    for i in range(len(data)):
        if re.search(pattern, data[i]):
            result = re.findall(pattern, data[i])
            Tax_ID_list = Tax_ID_list + result
    if len(Tax_ID_list) > 16:
        Tax_ID_list = []
        for i in range(len(data)):
            if re.search(pattern, data[i]):
                result = re.findall(r'^\s?GB\d{12}$', data[i])
                Tax_ID_list = Tax_ID_list + result
    if len(Tax_ID_list) == 1:
        return Tax_ID_list[0].replace(" ", "")
    else:
        return Tax_ID_list[1].replace(" ", "")


# 获取申报总货值 pdf_declarat_values
def get_good_value(data):
    for i in range(len(data)):
        if re.search(r'Place[  ]and[  ]date', data[i]):
            if len(data[i]) > 20:
                new_data = data[i][19:]
                if re.search(r'Total', new_data):
                    return new_data.split(" ")[1].replace(" ", "")
            else:
                if re.search(r'Total', data[i + 1]):
                    return data[i + 2].replace(" ", "")
                else:
                    return data[i + 1].split(" ")[1].replace(" ", "")


# 获取实际缴纳税金 pdf_Invoice_total
def get_Invoice_total(data):
    for i in range(len(data)):
        if re.search(r'CIF', data[i]):
            origin_data = data[i + 1]
            if '\xa0' in origin_data:
                first_data = origin_data.split("\xa0")[1]
                if '@' in first_data:
                    return first_data.split('@')[0][:-1]
                else:
                    return first_data
            if ' ' in origin_data:
                if is_number(data[i + 1].split(" ")[1]):
                    return data[i + 1].split(" ")[1].replace(" ", "")
                else:
                    if len(data[i + 1].split(" ")) > 2:
                        return data[i + 1].split(" ")[2].replace(" ", "")
                    else:
                        return data[i + 1].split(" ")[1].replace(" ", "")[4:]


# 判断是否可转数字
def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False


# 设置excel单元格颜色
def excel_set_color(path, sheet, row, column):
    # 根据路径打开excel
    wb = load_workbook(filename=path)
    ws = wb[sheet]
    # 填充单元格为黄色
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws[str(row) + str(column)].fill = fill
    wb.save(path)


# 获取最大行数
def get_max_row(ws):
    i = ws.max_row
    row_count = 0
    while ws.max_row > 0:
        row_dict = {i.value for i in ws[i]}
        if row_dict == {None}:
            i = i - 1
        else:
            row_count = i
            break
    return row_count


# 获取最大列数
def get_max_col(ws):
    i = ws.max_column
    col_count = 0
    while ws.max_column > 0:
        col_dict = {i.value for i in ws[i]}
        if col_dict == {None}:
            i = i - 1
        else:
            col_count = i
            break
    return col_count


# 获取excel_vat(1)
def get_vat_data(excel_path, sheet):
    wb = load_workbook(excel_path)
    ws = wb[sheet]
    excel_vat = []
    for i in range(1, get_max_row(ws) + 1):
        excel_vat.append(ws.cell(row=i, column=8).value)
    return excel_vat


# 不包税 获取vat对应的件数、申报价值(单条)
def get_vat_excel_data(excel_path, sheet, pdf_vat, excel_vat):
    wb = load_workbook(excel_path)
    ws = wb[sheet]
    for i in range(len(excel_vat)):
        if str(pdf_vat) == str(excel_vat[i]):
            row_num = i + 1
            row = []
            for col in range(1, 18):
                row.append(ws.cell(row=row_num, column=col).value)
            excel_shipment_num = row[1]  # 运单号
            excel_customer_num = row[2]  # 客户单号
            excel_customer_name = row[3]  # 客户名称
            excel_num = row[6]  # 件数
            excel_declarat_values = row[11]  # 申报价值
            return excel_shipment_num, excel_customer_num, excel_customer_name, pdf_vat, excel_num, excel_declarat_values, row_num


# 不包税 获取vat对应的件数、申报价值(单条&行数)
def get_vat_excel_data_other(excel_path, sheet, pdf_vat, excel_vat, row_num):
    try:
        wb = load_workbook(excel_path)
        ws = wb[sheet]
        row = []
        for col in range(1, 18):
            row.append(ws.cell(row=row_num, column=col).value)
        excel_shipment_num = row[1]  # 运单号
        excel_customer_num = row[2]  # 客户单号
        excel_customer_name = row[3]  # 客户名称
        excel_num = row[6]  # 件数
        excel_declarat_values = row[11]  # 申报价值
        return excel_shipment_num, excel_customer_num, excel_customer_name, pdf_vat, excel_num, excel_declarat_values, row_num
    except Exception as e:
        print(e)


# 不包税 获取vat对应的件数、申报价值(多条)
def get_vat_excel_data_list(excel_path, sheet, pdf_vat, excel_vat):
    wb = load_workbook(excel_path)
    ws = wb[sheet]
    row_nums = []
    excel_num_total = 0
    excel_declarat_values_total = float(0.00)
    for i in range(len(excel_vat)):
        if str(pdf_vat) == str(excel_vat[i]):
            row_num = i + 1
            row_nums.append(row_num)
            row = []
            for col in range(1, 18):
                row.append(ws.cell(row=row_num, column=col).value)
            excel_shipment_num = row[1]  # 运单号
            excel_customer_num = row[2]  # 客户单号
            excel_customer_name = row[3]  # 客户名称
            excel_num = row[6]  # 件数
            excel_num_total += int(excel_num)
            excel_declarat_values = row[11]  # 申报价值
            excel_declarat_values_total = excel_declarat_values_total + float(excel_declarat_values)
    return excel_shipment_num, excel_customer_num, excel_customer_name, pdf_vat, excel_num_total, excel_declarat_values_total, row_nums


# 不包税 获取vat对应的件数、申报价值(多条 特殊情况)
def get_vat_excel_data_list_new(excel_path, sheet, pdf_vat, excel_vat):
    try:
        wb = load_workbook(excel_path)
        ws = wb[sheet]
        row_nums = []
        excel_num_total = 0
        excel_declarat_values_total = float(0.00)
        if excel_vat != []:
            for i in range(len(excel_vat)):
                if str(pdf_vat) == str(excel_vat[i]):
                    row_num = i + 1
                    row = []
                    for col in range(1, 18):
                        row.append(ws.cell(row=row_num, column=col).value)
                    is_deal = row[16]
                    vat_num = row[7]
                    if vat_num == None:
                        continue
                    if is_deal == None:
                        row_nums.append(row_num)
                        excel_shipment_num = row[1]  # 运单号
                        excel_customer_num = row[2]  # 客户单号
                        excel_customer_name = row[3]  # 客户名称
                        excel_num = row[6]  # 件数
                        excel_num_total += int(excel_num)
                        excel_declarat_values = row[11]  # 申报价值
                        excel_declarat_values_total = excel_declarat_values_total + float(excel_declarat_values)
            return excel_shipment_num, excel_customer_num, excel_customer_name, pdf_vat, excel_num_total, excel_declarat_values_total, row_nums
        time.sleep(2)
    except Exception as e:
        print(e)
        time.sleep(2)


# 循环文件夹内文件夹
def get_folder(main_path):
    files = os.listdir(main_path)
    folder_paths = []
    for file in files:
        file_path = os.path.join(main_path, file)
        if os.path.isdir(file_path):
            folder_paths.append(file_path)
    # print(folder_paths)
    return folder_paths


# 循环文件夹内文件
def get_flie_name(folder_path):
    files = os.listdir(folder_path)
    file_paths = []
    for file in files:
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            if re.search(r'单号', file):
                continue
            file_paths.append(folder_path + "\\" + file)
    # print(file_paths)
    return file_paths


# 执行pdf操作
def pdf_run(file_path):
    data = get_pdf_data(file_path)
    # 件数
    pdf_nums = get_num(data).replace(" ", "")
    # 税号
    pdf_vat = get_Tax_ID(data)[0:11]
    # 申报价值
    pdf_declarat_values = get_Invoice_total(data)
    # 税金
    pdf_Invoice_total = get_good_value(data)
    return pdf_vat, pdf_nums, pdf_declarat_values, pdf_Invoice_total


# 判断是否存在vat
def is_vat_exict_second(pdf_vat, pdf_nums, pdf_declarat_values, pdf_Invoice_total, folder_path, remote_file):
    try:
        # excel文件
        excel_path = (folder_path + '\\单号.xlsx').replace("\\", "\\\\")
        # 获取excel_vat
        excel_vat = get_vat_data(excel_path, "Sheet1")
        if str(pdf_vat) in excel_vat:
            # 不包税 获取vat对应的件数、申报价值(单条)
            excel_data = get_vat_excel_data(excel_path, "Sheet1", pdf_vat,
                                            excel_vat)  # excel_shipment_num, excel_customer_num, excel_customer_name, pdf_vat, excel_num, excel_declarat_values, row_num
            # 不包税 判断vat、件数、申报价值 所有条件相等
            if str(pdf_vat) == str(pdf_vat) and int(pdf_nums) == int(excel_data[4]) and float(
                    pdf_declarat_values) == float(excel_data[5]):
                # is_vat_one_all_equal(pdf_vat, pdf_vat, pdf_nums, excel_data[4], pdf_declarat_values, excel_data[5], remote_file, folder_path, excel_data[6], pdf_Invoice_total)
                # print("不包税处理1")
                pass
            else:
                # 不包税 vat相等（多条）
                excel_data_list = get_vat_excel_data_list_new(excel_path, "Sheet1", pdf_vat,
                                                              excel_vat)  # excel_shipment_num, excel_customer_num, excel_customer_name, excel_vat, excel_num_total, excel_declarat_values_total, row_nums
                if excel_data_list != None:
                    # 循环行记录
                    for row_num in excel_data_list[6]:
                        excel_data_other = get_vat_excel_data_other(excel_path, "Sheet1", pdf_vat, excel_vat,
                                                                    row_num)  # excel_shipment_num, excel_customer_num, excel_customer_name, pdf_vat, excel_num, excel_declarat_values, row_num
                        # 检查是否存在vat、件数、申报价值 所有条件相等的记录
                        if str(excel_data_other[3]) == str(excel_data_other[3]) and int(pdf_nums) == int(
                                excel_data_other[4]) and float(pdf_declarat_values) == float(excel_data_other[5]):
                            tax_off(remote_file, folder_path, row_num, pdf_Invoice_total)
                    # 不包税 vat相等（多条）处理
                    is_vat_list_equal(pdf_vat, pdf_nums, excel_data_list[4], pdf_declarat_values, excel_data_list[5],
                                      remote_file, folder_path, excel_data_list[6], pdf_Invoice_total, excel_path)
                else:
                    pass
        else:
            # 包税处理
            # tax_up(pdf_vat, pdf_nums, pdf_declarat_values, pdf_Invoice_total, folder_path, remote_file, excel_path)
            # print("包税处理1")
            # 上一轮循环已经处理了包税情况，这一轮不处理包税情况，只针对不包税的合并情况处理
            pass
    except Exception as e:
        print(e)


# 不包税 判断vat、件数、申报价值 所有条件相等
def is_vat_one_all_equal(pdf_vat, excel_vat, pdf_nums, excel_nums, pdf_declarat_values, excel_declarat_values,
                         remote_file, folder_path, row, pdf_Invoice_total):
    if str(pdf_vat) == str(excel_vat) and int(pdf_nums) == int(excel_nums) and float(pdf_declarat_values) == float(
            excel_declarat_values):
        tax_off(remote_file, folder_path, row, pdf_Invoice_total)


# 不包税 vat相等（多条）
def is_vat_list_equal(pdf_vat, pdf_nums, excel_nums, pdf_declarat_values, excel_declarat_values, remote_file,
                      folder_path, row_nums, pdf_Invoice_total, excel_path):
    if int(pdf_nums) == int(excel_nums) and round(float(pdf_declarat_values), 2) == round(float(excel_declarat_values),
                                                                                          2):
        # 不包税处理
        tax_off_list(remote_file, folder_path, row_nums, pdf_Invoice_total, pdf_declarat_values)
        print("不包税处理2")
    else:
        if len(row_nums) == 1:
            # 包税处理
            tax_up(pdf_vat, pdf_nums, pdf_declarat_values, pdf_Invoice_total, folder_path, remote_file, excel_path)
            print("包税处理2")
        else:
            print("不处理")


# 不包税处理1
def tax_off(remote_file, folder_path, row_num, pdf_Invoice_total):
    try:
        excel_path = (folder_path + '\\单号.xlsx').replace("\\", "\\\\")
        # 写入excel
        wb = load_workbook(excel_path)
        # 读取excel原数据
        ws = wb['Sheet1']
        data = []
        for col in range(1, get_max_col(ws) + 1):
            data.append(ws.cell(row=row_num, column=col).value)
        # 获取提单号、运单号、客户名称、货值、税金、vat号
        lade_num = data[0]  # 提单号
        shipment_num = data[1]  # 运单号
        customer_num = data[2]  # 客户单号
        customer_name = data[3]  # 客户名称
        declarat_value = data[11]  # 申报价值
        vat_num = data[7]  # vat号
        sale_name = data[14]  # 销售代表
        finance_name = data[15]  # 财务代表
        # 标注 已处理
        ws.cell(row_num, 17, "已处理")
        wb.save(excel_path)
        # 写入不包税sheet
        ws = wb["不包税"]
        last_row = get_max_row(ws) + 1
        ws.cell(last_row, 1, lade_num)
        ws.cell(last_row, 2, shipment_num)
        ws.cell(last_row, 3, customer_name)
        ws.cell(last_row, 4, declarat_value)
        ws.cell(last_row, 8, pdf_Invoice_total)
        ws.cell(last_row, 17, vat_num)
        ws.cell(last_row, 18, sale_name)
        ws.cell(last_row, 19, finance_name)
        wb.save(excel_path)
        wb.close()
        # 重命名
        new_name = shipment_num + " " + customer_num + " " + customer_name + ".pdf"
        new_name = (folder_path + "\\" + new_name).replace("\\", "\\\\")
        if os.path.exists(new_name):
            pass
        os.rename(remote_file, new_name)
    except Exception as e:
        print(e)


# 不包税处理2
def tax_off_list(remote_file, folder_path, row_nums, pdf_Invoice_total, pdf_declarat_values):
    try:
        excel_path = (folder_path + '\\单号.xlsx').replace("\\", "\\\\")
        new_name = ""
        # 读取excel原数据
        for row_num in row_nums:
            # 写入excel
            wb = load_workbook(excel_path)
            # 读取excel原数据
            ws = wb['Sheet1']
            data = []
            for col in range(1, 18):
                data.append(ws.cell(row=row_num, column=col).value)
            # 获取提单号、运单号、客户名称、货值、税金、vat号
            lade_num = data[0]  # 提单号
            shipment_num = data[1]  # 运单号
            new_name += shipment_num + " "
            customer_num = data[2]  # 客户单号
            new_name += customer_num + " "
            customer_name = data[3]  # 客户名称
            declarat_value = data[11]  # 申报价值
            vat_num = data[7]  # vat号
            sale_name = data[14]  # 销售代表
            finance_name = data[15]  # 财务代表
            # 标注 已处理
            ws.cell(row_num, 17, "已处理")
            # 写入不包税sheet
            ws = wb["不包税"]
            last_row = get_max_row(ws) + 1
            ws.save(excel_path)
            wb.close()
            # 设置单元格颜色
            excel_set_color(excel_path, "不包税", "H", last_row)
            wb = load_workbook(excel_path)
            ws = wb['不包税']
            last_row = last_row
            ws.cell(last_row, 1, lade_num)
            ws.cell(last_row, 2, shipment_num)
            ws.cell(last_row, 3, customer_name)
            ws.cell(last_row, 4, declarat_value)
            ws.cell(last_row, 8, pdf_Invoice_total)
            ws.cell(last_row, 17, vat_num)
            ws.cell(last_row, 18, sale_name)
            ws.cell(last_row, 19, finance_name)
            wb.save(excel_path)
            wb.close()
        # 重命名
        new_name += customer_name
        if len(new_name) > 100:
            new_name = new_name[:99]
        new_name += ".pdf"
        new_name = (folder_path + "\\" + new_name).replace("\\", "\\\\")
        if os.path.exists(new_name):
            pass
        os.rename(remote_file, new_name)
    except Exception as e:
        print(e)


# 包税处理
def tax_up(pdf_vat, pdf_nums, pdf_declarat_values, pdf_Invoice_total, folder_path, remote_file, excel_path):
    try:
        wb = load_workbook(excel_path)
        ws = wb['Sheet1']
        # 获取提单号
        lade_num = ws.cell(row=2, column=1).value
        # 设置包税sheet信息
        ws = wb['包税']
        last_row = get_max_row(ws) + 1
        ws.cell(last_row, 1, lade_num)
        ws.cell(last_row, 2, pdf_vat)
        ws.cell(last_row, 3, pdf_nums)
        ws.cell(last_row, 4, pdf_Invoice_total)
        wb.save(excel_path)
        new_name = pdf_vat + " " + pdf_nums + " " + pdf_Invoice_total + ".pdf"
        new_name = (folder_path + "\\" + new_name).replace("\\", "\\\\")
        if os.path.exists(new_name):
            new_name = new_name.replace(".pdf", " 1.pdf")
        os.rename(remote_file, new_name)
        wb.close()
    except Exception as e:
        print(e)


# 二轮处理 两个税金单的vat相同，匹配两种情况（一条vat条件匹配，多条vat条件匹配）
def main_finally_deal():
    try:
        # 设置文件路径
        main_path = 'D:\\yingdaoFile\\demand3\\testFile'
        # 获取需要处理的提单文件夹
        folder_paths = get_folder(main_path)
        for i in range(len(folder_paths)):
            folder_path = folder_paths[i]
            # 获取需要处理的文件
            file_paths = get_flie_name(folder_path)
            kill_processes_by_names(['wps.exe'])
            for i in range(len(file_paths)):
                # 税金单文件
                pdf_path = file_paths[i]
                # 判断剩下的文件
                # if re.search(r'GB' ,pdf_path):
                if re.search(r'^(?=.*\bGB\b)(?!.*\b问题件\b).*', pdf_path):  # 修改判断，问题件存在需要再处理的数据
                    continue
                if re.search(r'UK', pdf_path):
                    continue
                # 获取税金单数据
                pdf_data = pdf_run(pdf_path)  # pdf_vat, pdf_nums, pdf_declarat_values, pdf_Invoice_total
                print(pdf_data)
                pdf_vat = pdf_data[0]  # vat单号
                pdf_nums = pdf_data[1]  # 件数
                pdf_declarat_values = pdf_data[2]  # 申报价值
                pdf_Invoice_total = pdf_data[3]  # 税金
                is_vat_exict_second(pdf_vat, pdf_nums, pdf_declarat_values, pdf_Invoice_total, folder_path, pdf_path)
    except Exception as e:
        print(e)
    finally:
        kill_processes_by_names(['wps.exe'])


if __name__ == '__main__':
    main_finally_deal()