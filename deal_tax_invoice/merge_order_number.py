import os, time, datetime, random, unicodedata, re, PyPDF2, psutil, openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from copy import deepcopy


# 杀死进程
def kill_processes_by_names(process_names):
    for proc in psutil.process_iter(['name']):
        try:
            if proc.info['name'] in process_names:
                proc.kill()
        except psutil.NoSuchProcess:
            print(f"进程 {proc.info['name']} 不存在.")
            continue
        except psutil.AccessDenied:
            print(f"无法终止进程 {proc.info['name']}.")


# 循环文件夹内文件夹
def get_folder(main_path):
    files = os.listdir(main_path)
    folder_paths = []
    for file in files:
        file_path = os.path.join(main_path, file)
        if os.path.isdir(file_path):
            folder_paths.append(file_path)
    # print(folder_paths)
    return folder_paths


# 循环文件夹内文件
def get_flie_name(folder_path):
    files = os.listdir(folder_path)
    file_paths = []
    for file in files:
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            if re.search(r'单号', file):
                file_paths.append(folder_path + "\\" + file)
    return file_paths


# 获取最大行数
def get_max_row(ws):
    i = ws.max_row
    row_count = 0
    while ws.max_row > 0:
        row_dict = {i.value for i in ws[i]}
        if row_dict == {None}:
            i = i - 1
        else:
            row_count = i
            break
    return row_count


# 获取最大列数
def get_max_col(ws):
    i = ws.max_column
    col_count = 0
    while ws.max_column > 0:
        col_dict = {i.value for i in ws[i]}
        if col_dict == {None}:
            i = i - 1
        else:
            col_count = i
            break
    return col_count


# 创建 单号(总).xlsx 表头信息
def create_excel_data(main_path):
    try:
        if not os.path.exists(main_path):
            Workbook().save(main_path)
        wb = load_workbook(main_path)
        # Sheet1
        ws = wb['Sheet']
        header_row = ["提单号", "运单号", "客户单号", "客户", "FBA单号", "集团单号", "件数",
                      "vat号", "服务", "交税方式", "报关方式", "申报价值", "收费重", "主品名", "销售代表", "财务代表",
                      "是否处理"]
        for i in range(0, len(header_row)):
            ws.cell(row=1, column=i + 1, value=header_row[i])
        # 包税sheet
        wb.create_sheet("包税", len(wb.sheetnames) + 1)
        ws = wb['包税']
        header_row = ["提单号", "VAT号", "件数", "税金", "分单号"]
        for i in range(0, len(header_row)):
            ws.cell(row=1, column=i + 1, value=header_row[i])
        # 不包税sheet
        wb.create_sheet("不包税", len(wb.sheetnames) + 1)
        ws = wb['不包税']
        header_row = ["提单号", "系统单号", "客户名称", "货值", "预收关税USD", "汇率", "预收关税RMB",
                      "实际关税英磅", "英磅汇率", "实际关税RMB", " ", "需退", "差额", "C88", "账单", "处理情况",
                      "VAT号", "销售代表", "财务代表"]
        for i in range(0, len(header_row)):
            ws.cell(row=1, column=i + 1, value=header_row[i])
        wb.save(main_path)
        wb.close()
    except Exception as e:
        print(e)


# 循环获取单号文件
def get_deal_file(main_path):
    try:
        # 获取需要处理的提单文件夹
        folder_paths = get_folder(main_path)
        excel_paths = []
        for i in range(len(folder_paths)):
            # 获取需要处理的文件
            folder_path = folder_paths[i]
            # 获取需要处理的文件
            excel_path = get_flie_name(folder_path)
            excel_paths += excel_path
        return excel_paths
    except Exception as e:
        print(e)


# excel复制粘贴操作
def copy_cell(copy_from, paste_to_cell):
    """复制粘贴某个区域
    :param copy_from 复制源 (应是 openpyxl 中的行或列)
    :param paste_to_cell 粘贴的左上角
    """
    # 记录边缘值
    for _copy_row in copy_from:  # 循环每一行
        for _row_cell in _copy_row:  # 循环每一列
            paste_to_cell.value = _row_cell.value
            # paste_to_cell._style = deepcopy(_row_cell.style)  # 复制样式
            paste_to_cell = paste_to_cell.offset(row=0, column=1)  # 右移1格
        paste_to_cell = paste_to_cell.offset(row=1, column=-len(_copy_row))  # 下移一行，列归位


# 循环sheet页 表格内容 （表头以下信息）
def deal_excel_data(old_path, old_sheet, new_path, new_sheet):
    try:
        # 获取复制excel的最大值
        old_wb = load_workbook(old_path)
        old_ws = old_wb[old_sheet]
        old_row_count = get_max_row(old_ws)
        old_col_count = get_max_col(old_ws)

        new_wb = load_workbook(new_path)
        new_ws = new_wb[new_sheet]
        new_max_row = get_max_row(new_ws) + 1

        # 使用 iter_rows 获取需要复制的单元格区域
        copy_from = old_ws.iter_rows(min_row=2, max_row=old_row_count, min_col=1, max_col=20)
        # 获取目标单元格
        paste_to_cell = new_ws['A' + str(new_max_row)]

        # 复制内容
        copy_cell(copy_from, paste_to_cell)
        new_wb.save(new_path)
    except Exception as e:
        print(e)


# 汇总数据
def main_merge_order():
    try:
        kill_processes_by_names(['wps.exe'])
        time.sleep(1)
        main_path = 'D:\\yingdaoFile\\demand3\\testFile'
        main_excel_path = main_path + '\\汇总单号.xlsx'
        create_excel_data(main_excel_path)
        time.sleep(1)
        kill_processes_by_names(['wps.exe'])
        excel_paths = get_deal_file(main_path)
        # 循环处理单号.xlsx对应的sheet
        for excel_data in excel_paths:
            # 复制Sheet1
            deal_excel_data(excel_data, 'Sheet1', main_excel_path, 'Sheet')
            # # 复制包税
            deal_excel_data(excel_data, '包税', main_excel_path, '包税')
            # # 复制不包税
            deal_excel_data(excel_data, '不包税', main_excel_path, '不包税')
    except Exception as e:
        print(e)


if __name__ == '__main__':
    main_merge_order()