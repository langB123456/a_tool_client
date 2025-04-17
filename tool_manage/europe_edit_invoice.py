import os
import re
import time
import zipfile
from zipfile import ZipFile
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
import requests
import argostranslate.translate as translate
import argostranslate.package as package


# 初始化模型
def initial_model():
    # 更新可用模型列表
    package.update_package_index()

    # 获取中文到英文的模型
    available_packages = package.get_available_packages()
    package_to_install = next(
        filter(
            lambda x: x.from_code == "zh" and x.to_code == "en",
            available_packages
        )
    )

    # 下载并安装模型
    package.install_from_path(package_to_install.download())


# 获取最大行数
def get_max_row(ws):
    i = ws.max_row
    row_count = 0
    while ws.max_row > 0:
        row_dict = {i.value for i in ws[i]}
        if row_dict == {None}:
            i = i - 1
        else:
            row_count = i
            break
    return row_count


# 获取最大列数
def get_max_col(ws):
    i = ws.max_column
    col_count = 0
    while ws.max_column > 0:
        col_dict = {i.value for i in ws[i]}
        if col_dict == {None}:
            i = i - 1
        else:
            col_count = i
            break
    return col_count


# 模糊查询文件名称
def get_file_path(folder_path: str, file_name: str):
    files = os.listdir(folder_path)
    for file in files:
        if file_name in file:
            file_path = os.path.join(folder_path, file)
            return file_path


# 登录天图
# base_url = https://test.tttxex.com/
def get_token(base_url, user, pwd):
    url = base_url + "/api/sys/auth/login"
    headers = {"content-type": "application/json; charset=utf-8"}
    data = {"password": pwd,
            "username": user}
    response = requests.post(url, headers=headers, json=data)
    token = response.json()["data"]["token"]
    return token


# 查询提单id 柜号 cabinet_no_list
def get_lade_id_by_cabinet_no(base_url, token, cabinet_num):
    url = base_url + "/api/ladeorder/lade_order_list"
    headers = {"content-type": "application/json; charset=utf-8",
               "Token": token}
    data = {
        "code": "",
        "customer_id": [],
        "channel_list": [],
        "country_code_list": [],
        "lade_type": None,
        "declare_id": "",
        "customs_id": "",
        "send_site": "",
        "to_site": "",
        "bloc_no": "",
        "ltd_id_list": [],
        "page": 1,
        "limit": 100,
        "status": 0,
        "recipient_info_list": [],
        "cabinet_no_list": [cabinet_num]
    }
    response = requests.post(url, headers=headers, json=data)
    lade_id = response.json()["data"]["list"][0]["id"]
    if lade_id:
        return lade_id


# 查询提单id 提单号 lade_number_list
def get_lade_id_by_lade_num(base_url, token, lade_num):
    url = base_url + "/api/ladeorder/lade_order_list"
    headers = {"content-type": "application/json; charset=utf-8",
               "Token": token}
    data = {
        "code": "",
        "customer_id": [],
        "channel_list": [],
        "country_code_list": [],
        "lade_type": None,
        "declare_id": "",
        "customs_id": "",
        "send_site": "",
        "to_site": "",
        "bloc_no": "",
        "ltd_id_list": [],
        "page": 1,
        "limit": 100,
        "status": 2,
        "recipient_info_list": [],
        "cabinet_no_list": "",
        "lade_number_list": [lade_num]
    }
    response = requests.post(url, headers=headers, json=data)
    lade_id = response.json()["data"]["list"][0]["id"]
    if lade_id:
        return lade_id


# 下载天图的zip文件
def download_zip(base_url, token, lade_id, export_type, file_path):
    try:
        url = base_url + "/api/ladeorder/exportByTemplate"
        headers = {"content-type": "application/json",
                   "token": token}
        data = {"exportTypes": [export_type], "idList": [lade_id]}
        response = requests.post(url, headers=headers, json=data)
        response.raise_for_status()
        with open(file_path, 'wb') as f:
            f.write(response.content)
    except Exception as e:
        print(e)


# 解压包
def unzip_file(zip_path, extract_to):
    # 打开 ZIP 文件并解压
    with ZipFile(zip_path, 'r') as zip_ref:
        files = zip_ref.namelist()
        zip_ref.extractall(extract_to)
    os.remove(zip_path)
    return (extract_to + "\\" + files[0]).replace("\\", "\\\\")


# 压缩包
def zip_folder(folder_path, output_path):
    try:
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # os.walk() 生成文件夹中的文件名和子文件夹名
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    # 获取文件相对路径
                    file_path = os.path.join(root, file)
                    # 在 ZIP 文件中添加文件，并将路径设置为相对路径
                    zipf.write(str(file_path), os.path.relpath(str(file_path), folder_path))
        print("zip successful")
    except Exception as e:
        print(e)


# 循环文件夹内文件
def get_file_name(folder_path):
    files = os.listdir(folder_path)
    file_paths = []
    for file in files:
        file_path = os.path.join(folder_path, file)
        file_paths.append(folder_path + "\\" + file)
    return file_paths


# 查询excel关键字
def select_key_name(ws, key_name):
    for row in range(1, 30):
        for col in range(1, 30):
            if key_name == ws.cell(row=row, column=col).value:
                return row, col


# 读取excel
def read_excel_col(excel_path, sheet_name, col_key):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    col_keys = []
    col_count = get_max_row(ws)
    for i in range(2, col_count + 1):
        col_keys.append(ws.cell(row=i, column=col_key).value)
    return col_keys


# 清空文件夹的文件
def clear_files(folder_path):
    files = get_file_name(folder_path)
    if files != []:
        for file in files:
            os.remove(file)


# 系统下载-海运and铁路
def download_sys_file_first(first_list, second_list, right_key, filename, export_types):
    base_url = "https://fba.tttxex.com/"
    folder = "D:\\yingdaoFile\\demand8\\exportFile\\"
    file_path = folder + filename + '.zip'
    token = get_token(base_url, "admin", "otvxQyE3+uqbha16nvMt1g==")
    final_list = []
    if len(first_list) == len(second_list):
        for i in range(len(first_list)):
            if right_key == first_list[i]:
                final_list.append(second_list[i])
    print(final_list)
    # 解压文件
    lade_id = get_lade_id_by_cabinet_no(base_url, token, final_list[0])
    print(lade_id)
    download_zip(base_url, token, lade_id, export_types, file_path)
    first = unzip_file(file_path, folder)
    second = unzip_file(first, folder)
    return first


# 系统下载-空运
def download_sys_file_second(first_list, second_list, right_key, filename, export_types):
    base_url = "https://fba.tttxex.com/"
    folder = "D:\\yingdaoFile\\demand8\\exportFile\\"
    file_path = folder + filename + '.zip'
    token = get_token(base_url, "admin", "otvxQyE3+uqbha16nvMt1g==")
    final_list = []
    if len(first_list) == len(second_list):
        for i in range(len(first_list)):
            if right_key == first_list[i]:
                final_list.append(second_list[i])
    print(final_list)
    # 解压文件
    lade_id = get_lade_id_by_lade_num(base_url, token, final_list[0])
    print(lade_id)
    download_zip(base_url, token, lade_id, export_types, file_path)
    first = unzip_file(file_path, folder)
    second = unzip_file(first, folder)
    return first


# 中译英 翻译
def translate_tool(excel_path, chinese_text):
    def is_chinese(text):
        """检查文本是否包含中文字符"""
        return any('\u4e00' <= char <= '\u9fff' for char in text)

    # 确保输入是字符串
    if not isinstance(chinese_text, str):
        print(f"{excel_path} 的输入 {chinese_text} 不是字符串，返回原值")
        return str(chinese_text)

    # 尝试翻译
    for attempt in range(6):  # 初始尝试 + 5 次重试
        # 调用谷歌的包，需要外网
        # try:
        #     translator = GoogleTranslator(source="zh-CN", target="en")
        #     new_text = translator.translate(chinese_text)
        #
        #     # 验证翻译结果：不含中文字符
        #     if new_text and not is_chinese(new_text):
        #         if attempt > 0:
        #             print(f"{excel_path} 第 {attempt} 次重试翻译成功: {chinese_text} -> {new_text}")
        #         return new_text
        #     else:
        #         print(f"{excel_path} 翻译结果仍包含中文或无效: {new_text}, 第 {attempt + 1} 次重试")
        #
        # except Exception as e:
        #     print(f"{excel_path} 翻译失败: {chinese_text}, 错误: {e}, 第 {attempt + 1} 次尝试")
        #     if attempt < 5:  # 最多重试 5 次
        #         time.sleep(1)  # 延迟 1 秒，避免频繁请求
        #     continue

        # 调用本地模型包，需要初始化模型
        try:
            new_text = translate.translate(chinese_text, from_code="zh", to_code="en")

            # 验证翻译结果：不含中文字符
            if new_text and not is_chinese(new_text):
                if attempt > 0:
                    print(f"{excel_path} 第 {attempt} 次重试翻译成功: {chinese_text} -> {new_text}")
                return new_text
            else:
                print(f"{excel_path} 翻译结果仍包含中文或无效: {new_text}, 第 {attempt + 1} 次重试")

        except Exception as e:
            print(f"{excel_path} 翻译失败: {chinese_text}, 错误: {e}, 第 {attempt + 1} 次尝试")
            if attempt < 5:  # 最多重试 5 次
                time.sleep(1)  # 延迟 1 秒，避免频繁请求
            continue

    # 所有尝试失败
    print(f"{excel_path} 的文本 {chinese_text} 翻译失败 5 次，请手动修改")
    return chinese_text  # 返回原文本作为默认值


# 特殊处理excel，针对某列中译英
def read_special_excel(excel_path, sheet_name, key_name):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    row, col = select_key_name(ws, key_name)
    row_max = get_max_row(ws)
    row_end: int
    # 获取需要操作的行
    for i in range(row, row_max):
        value = ws.cell(row=i, column=col).value
        if value is None:
            row_end = i
            break
    # 操作列表行
    for i in range(row + 1, row_end):
        value = ws.cell(row=i, column=col).value
        if value:
            # 判断纯中文
            if all('\u4e00' <= char <= '\u9fff' for char in value):
                # 转换英文
                print(value, i, col)
                value = translate_tool(excel_path, value)
                ws.cell(row=i, column=col, value=value)
            else:
                # 删除中文
                value = re.sub(r'[\u4e00-\u9fff]', '', value).replace("/", "")
                ws.cell(row=i, column=col, value=value)
    ws.row_dimensions[row + 1].height = 15
    wb.save(excel_path)
    wb.close()


# 处理品牌
def read_brand_excel(excel_path, sheet_name, key_name):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    row, col = select_key_name(ws, key_name)
    row_max = get_max_row(ws)
    row_end: int
    # 获取需要操作的行
    for i in range(row, row_max):
        value = ws.cell(row=i, column=col).value
        if value is None:
            row_end = i
            break
    # 操作列表行
    for i in range(row + 1, row_end):
        value = ws.cell(row=i, column=col).value
        # 判断纯英文
        if re.match(r'^[a-zA-Z0-9\s.,!?\'"-]*$', value):
            # 转换英文
            pass
        else:
            ws.cell(row=i, column=col, value="/")
    ws.row_dimensions[row + 1].height = 15
    wb.save(excel_path)


# 主流程
def main_flow(export_folder, excel_path, e_sheet_name, first_key: int, second_key: int, right_key, template_name,
              x_sheet_name, key_name):
    clear_files(export_folder)
    folder = "D:\\yingdaoFile\\demand8\\exportFile"
    col_c = read_excel_col(excel_path, e_sheet_name, first_key)
    col_i = read_excel_col(excel_path, e_sheet_name, second_key)
    # 系统下载
    rename_file = download_sys_file_first(col_c, col_i, right_key, template_name)
    files = get_file_name(export_folder)
    for file in files:
        print(file)
        read_special_excel(file, x_sheet_name, key_name)
    zip_folder(folder, str(rename_file).replace("exportFile\\", ""))
    # zip_folder(folder, "D:\yingdaoFile\demand8\exportFile\\1.UKSZ202412260382.xlsx")


# 海运:代理【K】--欧线-模板一
def sea_transport_model_first(export_folder, excel_path, e_sheet_name, template_name):
    clear_files(export_folder)
    right_key = "K"
    x_sheet_name = "箱单发票"
    first_col_key = "Model"
    second_col_key = "Material"
    third_col_key = "for"
    export_types = "O_LINE_ONE"
    folder = "D:\\yingdaoFile\\demand8\\exportFile"
    col_c = read_excel_col(excel_path, e_sheet_name, 2)
    col_i = read_excel_col(excel_path, e_sheet_name, 1)
    # 系统下载
    rename_file = download_sys_file_first(col_c, col_i, right_key, template_name, export_types)
    files = get_file_name(export_folder)
    for file in files:
        print(file)
        read_brand_excel(file, x_sheet_name, first_col_key)
        read_special_excel(file, x_sheet_name, second_col_key)
        read_special_excel(file, x_sheet_name, third_col_key)
    zip_folder(folder, str(rename_file).replace("exportFile\\", ""))


# 海运：代理【Y】--欧线-模板五-海运二
def sea_transport_model_second(export_folder, excel_path, e_sheet_name, template_name):
    clear_files(export_folder)
    right_key = "Y"
    x_sheet_name = "INVOICE"
    first_col_key = "Material"
    export_types = "O_LINE_FIVE_TWO"
    folder = "D:\\yingdaoFile\\demand8\\exportFile"
    col_c = read_excel_col(excel_path, e_sheet_name, 2)
    col_i = read_excel_col(excel_path, e_sheet_name, 1)
    # 系统下载
    rename_file = download_sys_file_first(col_c, col_i, right_key, template_name, export_types)
    files = get_file_name(export_folder)
    for file in files:
        print(file)
        read_special_excel(file, x_sheet_name, first_col_key)
    zip_folder(folder, str(rename_file).replace("exportFile\\", ""))


# 空运：代理【Y】欧线-模板四-空运
def sky_transport_model_fourth(export_folder, excel_path, e_sheet_name, template_name):
    clear_files(export_folder)
    right_key = "Y"
    x_sheet_name = "INVOICE"
    first_col_key = "Material"
    export_types = "O_LINE_FOUR"
    folder = "D:\\yingdaoFile\\demand8\\exportFile"
    col_c = read_excel_col(excel_path, e_sheet_name, 2)
    col_i = read_excel_col(excel_path, e_sheet_name, 1)
    # 系统下载
    rename_file = download_sys_file_second(col_c, col_i, right_key, template_name, export_types)
    files = get_file_name(export_folder)
    for file in files:
        print(file)
        read_special_excel(file, x_sheet_name, first_col_key)
    zip_folder(folder, str(rename_file).replace("exportFile\\", ""))


# 空运：代理【IT】英国清关资料 发票 模板七-(ITW)
def sky_transport_model_seventh(export_folder, excel_path, e_sheet_name, template_name):
    clear_files(export_folder)
    right_key = "IT"
    x_sheet_name = "Commercial Invoice"
    first_col_key = "material"
    export_types = "UK_CUSTOMS_CLEARANCE_SEVEN_ITW"
    folder = "D:\\yingdaoFile\\demand8\\exportFile"
    col_c = read_excel_col(excel_path, e_sheet_name, 2)
    col_i = read_excel_col(excel_path, e_sheet_name, 1)
    # 系统下载
    rename_file = download_sys_file_second(col_c, col_i, right_key, template_name, export_types)
    files = get_file_name(export_folder)
    for file in files:
        if "Invoice" in file:
            print(file)
            read_special_excel(file, x_sheet_name, first_col_key)
    zip_folder(folder, str(rename_file).replace("exportFile\\", ""))


# 铁卡：代理【IT】英国清关资料 发票 模板七-(ITW)
def iron_car_transport_model_fourth(export_folder, excel_path, e_sheet_name, template_name):
    clear_files(export_folder)
    right_key = "IT"
    x_sheet_name = "Commercial Invoice"
    first_col_key = "material"
    export_types = "UK_CUSTOMS_CLEARANCE_SEVEN_ITW"
    folder = "D:\\yingdaoFile\\demand8\\exportFile"
    col_c = read_excel_col(excel_path, e_sheet_name, 2)
    col_i = read_excel_col(excel_path, e_sheet_name, 1)
    # 系统下载
    rename_file = download_sys_file_second(col_c, col_i, right_key, template_name, export_types)
    files = get_file_name(export_folder)
    for file in files:
        if "Invoice" in file:
            print(file)
            read_special_excel(file, x_sheet_name, first_col_key)
    zip_folder(folder, str(rename_file).replace("exportFile\\", ""))


# 铁卡：代理【K】欧线-模板一
def iron_car_transport_model_first(export_folder, excel_path, e_sheet_name, template_name):
    clear_files(export_folder)
    right_key = "K"
    x_sheet_name = "箱单发票"
    first_col_key = "Model"
    second_col_key = "Material"
    third_col_key = "for"
    export_types = "O_LINE_ONE"
    folder = "D:\\yingdaoFile\\demand8\\exportFile"
    col_c = read_excel_col(excel_path, e_sheet_name, 2)
    col_i = read_excel_col(excel_path, e_sheet_name, 1)
    # 系统下载
    rename_file = download_sys_file_second(col_c, col_i, right_key, template_name, export_types)
    files = get_file_name(export_folder)
    for file in files:
        print(file)
        read_brand_excel(file, x_sheet_name, first_col_key)
        read_special_excel(file, x_sheet_name, second_col_key)
        read_special_excel(file, x_sheet_name, third_col_key)
    zip_folder(folder, str(rename_file).replace("exportFile\\", ""))


if __name__ == '__main__':
    initial_model()
    export_folder = "D:\\yingdaoFile\\demand8\\exportFile\\"
    excel_path = "D:\\yingdaoFile\\demand8\\欧线导出配置.xlsx"
    sea_transport_model_first(export_folder, excel_path, "海运", "欧线-模板一")
    sea_transport_model_second(export_folder, excel_path, "海运", "欧线-模板五-海运二")
    sky_transport_model_fourth(export_folder, excel_path, "空运", "欧线-模板四-空运")
    sky_transport_model_seventh(export_folder, excel_path, "空运", "英国清关资料 发票 模板七-(ITW)")
    iron_car_transport_model_fourth(export_folder, excel_path, "铁卡", "英国清关资料 发票 模板七-(ITW)")
    iron_car_transport_model_first(export_folder, excel_path, "铁卡", "欧线-模板一")
