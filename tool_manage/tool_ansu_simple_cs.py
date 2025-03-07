import requests
import re
import time
import openpyxl
from datetime import datetime
import keyboard
import uiautomation as auto
import pyperclip
import xlrd
from PIL import Image
import pytesseract, os, pyautogui, cv2, math, random, string, subprocess
import numpy as np
import subprocess
import tkinter as tk
import common.use_process

# 登录安速
def get_token(base_url, user, pwd):
    url = base_url + "/api/base/biz/user/login?clientType=MANAGE"
    headers = {"content-type": "application/json; charset=utf-8"}
    data = {"password": pwd,
            "username": user,
            "clientType": "MANAGE"}
    response = requests.post(url, headers=headers, json=data)
    token = response.json()["data"]
    # print(token)
    return token


# 数据导出
def get_data(base_url, authorization, shipment_nums, file_name):
    url = base_url + "api/base/biz/bizPreOrderShipment/manager/selectBypage"
    headers = {"content-type": "application/json", "authorization": authorization}
    data = {"mergeNoList": [], "waybillNumList": shipment_nums, "keyWordsList": [], "fbaNumberList": [],
            "transferOrderNumberList": [], "bizIdList": [], "followIdList": [], "channelIdList": [],
            "warehouseCodeList": [], "warehousePropertyList": [], "remark": "", "markList": [], "customerIdList": [],
            "zipCode": "", "countryRegionIdList": [], "customsTaxesTypeList": [], "innerRemark": "", "pageNum": 1,
            "pageSize": 50, "companyIds": [], "outWarehouseBrokerIds": [], "warehouseNameList": [], "beforePageNum": 0,
            "statusList": [-2],
            "saPageInfo": {"total": "0", "list": [], "searchAfter": None, "pageSize": 50, "pages": 0,
                           "nextSearchAfter": None}}
    response = requests.post(url, headers=headers, json=data)
    for i in range(len(response.json()["data"]["saPageInfo"]['list'])):
        shipment_id = response.json()["data"]["saPageInfo"]['list'][i]['waybillNum']
        print(shipment_id)


# 获取导出模板列表
def get_export_template_id(base_url, authorization):
    url = base_url + '/api/export/biz/exportTemplate/page'
    headers = {"content-type": "application/json", "authorization": authorization}
    data = {"pageNum": 0, "pageSize": 999, "total": 0, "beforePageNum": 0, "name": "", "status": 1, "dataScope": 0}
    response = requests.post(url, headers=headers, json=data)
    for i in range(len(response.json()['data']['records'])):
        name = response.json()['data']['records'][i]['name']
        if name == '安速-客服-数据发送':
            id = response.json()['data']['records'][i]['id']
            print(id)
            return id


# 导出操作
def export_file(base_url, authorization, shipment_nums, biz_export_template_id):
    url = base_url + 'api/export/biz/exportTemplate/customerCreateTemplateFile'
    headers = {"content-type": "application/json", "authorization": authorization}
    data = {
      "mergeNoList": [],
      "waybillNumList": shipment_nums,
      "keyWordsList": [],
      "fbaNumberList": [],
      "transferOrderNumberList": [],
      "bizIdList": [],
      "followIdList": [],
      "channelIdList": [],
      "warehouseCodeList": [],
      "warehousePropertyList": [],
      "remark": "",
      "markList": [],
      "customerIdList": [],
      "zipCode": "",
      "countryRegionIdList": [],
      "customsTaxesTypeList": [],
      "innerRemark": "",
      "pageNum": 1,
      "pageSize": 50,
      "companyIds": [],
      "outWarehouseBrokerIds": [],
      "warehouseNameList": [],
      "beforePageNum": 0,
      "statusList": [
        -2
      ],
      "bizExportTemplateId": biz_export_template_id,
      "selectStatus": -2
    }
    response = requests.post(url, headers=headers, json=data)
    if response.json()['data'] == '模版创建成功':
        return True
    else:
        return False


# 获取消息列表 下载地址
def get_news_list(base_url, authorization, export_name):
    url = base_url + 'api/base/biz/messageNotice/getBizMessageNotice'
    headers = {"content-type": "application/json", "authorization": authorization}
    response = requests.post(url, headers=headers)
    for i in range(len(response.json()['data'])):
        name = response.json()['data'][i]['name']
        if name == export_name:  # '安速-客服-数据发送'
            operate = response.json()['data'][i]['operate']
            return operate


# 下载文件
def download_file(url, file_path, file_name):
    print(file_path + file_name)
    if not os.path.exists(file_path + file_name):
        response = requests.get(url)
        with open(file_path + file_name, 'wb') as f:
            f.write(response.content)


# 获取最大行数
def get_max_row(ws):
    i = ws.max_row
    row_count = 0
    while ws.max_row > 0:
        row_dict = {i.value for i in ws[i]}
        if row_dict == {None}:
            i = i - 1
        else:
            row_count = i
            break
    return row_count


# 读取excel 获取运单号
def get_shipment_num_list(excel_path, sheet):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet]
    shipment_num_list = []
    row_count = get_max_row(ws)
    for i in range(2, row_count+1):
        shipment_num_list.append(ws.cell(row=i, column=1).value)
    return shipment_num_list


# 安速下载附件操作
def system_download(base_url, user, pwd):
    authorization = get_token(base_url, user, pwd)
    time.sleep(0.5)
    shipment_nums = get_shipment_num_list('D:\\group_send\\ansu_deal\\运单号.xlsx', 'Sheet1')
    biz_export_template_id = get_export_template_id(base_url, authorization)
    if export_file(base_url, authorization, shipment_nums, biz_export_template_id):
        time.sleep(0.5)
        download_url = get_news_list(base_url, authorization, '安速-客服-数据发送')
        time.sleep(0.5)
        download_file(download_url, 'D:\\group_send\\ansu_deal\\', '安速-客服-数据发送.xlsx')


# 过滤excel重复值
def skip_all_duplicates(arr):
    seen = []
    for value in arr:
        if value in seen:
            continue  # 如果当前值已经在集合中，则跳过
        seen.append(value)
    return seen


# 业务员匹配群名称处理
def sales_match_group():
    deal_path = 'D:\\group_send\\ansu_deal\\安速-客服-数据发送.xlsx'
    read_wb = openpyxl.load_workbook('D:\\group_send\\ansu_deal\\安速客服群发-基础信息维护.xlsx')
    read_ws = read_wb['群名称']
    row_count = get_max_row(read_ws)
    deal_wb = openpyxl.load_workbook(deal_path)
    deal_ws = deal_wb['Sheet1']
    deal_row_count = get_max_row(deal_ws)
    for i in range(2, deal_row_count+1):
        deal_sales_name = deal_ws.cell(row=i, column=3).value
        for j in range(1, row_count + 1):
            if deal_sales_name == read_ws.cell(row=j, column=1).value:
                deal_ws.cell(row=i, column=4, value=read_ws.cell(row=j, column=2).value)
                deal_ws.cell(row=i, column=5, value=read_ws.cell(row=j, column=3).value)
    if '运单对应群名' not in deal_wb.sheetnames:
        deal_wb.create_sheet('运单对应群名')
    other_ws = deal_wb['运单对应群名']
    head_row = ['客户简称', '运单号', '业务员名称', '群聊名称', '发送结果']
    for i in range(len(head_row)):
        other_ws.cell(row=1, column=i+1, value=head_row[i])
        i += 1
    # 获取唯一的群名称
    simple_sales_name = []
    for i in range(2, deal_row_count+1):
        deal_sales_name = deal_ws.cell(row=i, column=4).value
        simple_sales_name.append(deal_sales_name)
    select_data = skip_all_duplicates(simple_sales_name)
    k = 1
    # 根据唯一群名进行排序
    for i in range(len(select_data)):
        for j in range(2, deal_row_count+1):
            if select_data[i] == deal_ws.cell(row=j, column=4).value and deal_ws.cell(row=j, column=4).value is not None:
                k += 1
                other_ws.cell(row=k, column=1, value=deal_ws.cell(row=j, column=1).value)
                other_ws.cell(row=k, column=2, value=deal_ws.cell(row=j, column=2).value)
                other_ws.cell(row=k, column=3, value=deal_ws.cell(row=j, column=3).value)
                other_ws.cell(row=k, column=4, value=deal_ws.cell(row=j, column=4).value)
    # 合并单元格
    group_column = 'D'
    send_result_column = 'E'
    cells = [cell for cell in other_ws[group_column] if cell.row > 1]
    start_row = None
    current_value = None
    for idx, cell in enumerate(cells, start=2):
        value = cell.value
        if value == current_value:
            continue
        else:
            if start_row is not None and start_row != idx - 1:
                other_ws.merge_cells(f'{group_column}{start_row}:{group_column}{idx-1}')
                other_ws.merge_cells(f'{send_result_column}{start_row}:{send_result_column}{idx - 1}')
            current_value = value
            start_row = idx
    if start_row is not None and start_row != len(cells) + 1:
        other_ws.merge_cells(f'{group_column}{start_row}:{group_column}{len(cells+ 1)}')
        other_ws.merge_cells(f'{send_result_column}{start_row}:{send_result_column}{len(cells + 1)}')
    if 'Sheet3' not in deal_wb.sheetnames:
        deal_wb.create_sheet('Sheet3')
    final_ws = deal_wb['Sheet3']
    head_row = ['公司名称', '群聊名称', '发送结果']
    for i in range(len(head_row)):
        final_ws.cell(row=1, column=i + 1, value=head_row[i])
    for i in range(len(select_data)):
        if select_data[i] is not None:
            final_ws.cell(row=i+2, column=2, value=select_data[i])
    deal_wb.save(deal_path)
    deal_wb.close()


def get_photo_position(image_path):
    try:
        image = pyautogui.locateOnScreen(image_path, grayscale=True)
        center_x, center_y = pyautogui.center(image)
        return center_x, center_y
    except Exception as e:
        print(e, image_path)


# 复制文本信息
def copy_text_file_to_clipboard(file_path):
    with open(file_path, "r", encoding='utf8') as file:
        text = file.read()
        if text is not None:
            pyperclip.copy(text)


# 复制文件
def copy_file_to_clipboard(file_path):
    with open(file_path, "rb") as file:
        image = file.read()
        pyperclip.copy(image)


# 查询发送人信息
def search_sender(search_path, clear_path):
    user_name = '方天研发部-影刀(影刀)'
    # 搜索输入框查询
    search_x, search_y = get_photo_position(search_path)
    pyautogui.moveTo(search_x, search_y)
    pyautogui.click()
    keyboard.write(user_name)
    # keyboard.press_and_release('enter')
    time.sleep(1)
    pyautogui.moveTo(search_x + 70, search_y + 70)
    pyautogui.click()
    time.sleep(0.5)
    # 清空消息记录
    pyautogui.rightClick(search_x + 60, search_y + 60)
    clear_x, clear_y = get_photo_position(clear_path)
    pyautogui.moveTo(clear_x, clear_y)
    pyautogui.click()
    second_confirm_x, second_confirm_y = get_photo_position('D:\\group_send\\picture_manage\\second_confirm.png')
    pyautogui.moveTo(second_confirm_x, second_confirm_y)
    pyautogui.click()


def new_search_sender(user_name, search_path):
    # 搜索输入框查询
    search_x, search_y = get_photo_position(search_path)
    pyautogui.moveTo(search_x, search_y)
    pyautogui.click()
    keyboard.write(user_name)
    # keyboard.press_and_release('enter')
    time.sleep(1)
    pyautogui.moveTo(search_x + 70, search_y + 70)
    pyautogui.click()


# 点击逐条转发按钮
def forward_item(file_path):
    forward_item_x, forward_item_y = get_photo_position(file_path)
    pyautogui.moveTo(forward_item_x, forward_item_y)
    pyautogui.click()


def position_and_move(file_path):
    send_button_x, send_button_y = get_photo_position(file_path)
    pyautogui.moveTo(send_button_x, send_button_y)
    pyautogui.click()


# 发送按钮
def send_button(file_path):
    send_button_x, send_button_y = get_photo_position(file_path)
    # pyautogui.moveTo(send_button_x, send_button_y)
    pyautogui.moveTo(send_button_x - 150, send_button_y)
    pyautogui.click()


# 发送文本信息
def send_text(file_path):
    copy_text_file_to_clipboard(file_path)
    time.sleep(0.5)
    keyboard.press_and_release('ctrl+v')
    keyboard.press_and_release('enter')


def new_send_text(file_path, group_name):
    copy_text_file_to_clipboard(file_path)
    time.sleep(0.5)
    keyboard.press_and_release('ctrl+v')
    time.sleep(0.5)
    text = '\n' + get_customer_and_shipment(group_name)
    pyperclip.copy(text)
    time.sleep(0.5)
    keyboard.press_and_release('ctrl+v')
    keyboard.press_and_release('enter')


# 循环文件夹内文件 消息内容.txt 除外
def get_file_name(folder_path):
    files = os.listdir(folder_path)
    file_paths = []
    for file in files:
        if re.search('消息内容', file):
            continue
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            file_paths.append(folder_path + "\\" + file)
    # print(file_paths)
    return file_paths


# 发送文件信息
def send_file(file_path):
    if file_path is not None:
        # copy_file_to_clipboard(file_path)
        args = ['powershell', 'Get-Item {} | Set-Clipboard'.format(file_path)]
        subprocess.Popen(args=args, encoding='utf-8')
        time.sleep(0.5)
        keyboard.press_and_release('ctrl+v')
        time.sleep(0.5)


# 输入信息
def send_point(text):
    keyboard.write(text)
    keyboard.press_and_release('enter')


# 生成随机值
def generate_random_letters(length=5):
    # 选择所有可能的字母（包括大写和小写字母）
    letters = string.ascii_letters.lower()  # 包含 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # 使用 random.choices 随机选择指定数量的字母
    random_letters = ''.join(random.choices(letters, k=length))
    new_letters = 'PillowPillow' + random_letters + 'PillowPillow'
    return new_letters


# 处理图片
def deal_picture(image_path, text_data):
    image = Image.open(image_path)
    pytesseract.load_freq_dawg = False
    pytesseract.load_system_dawg = False
    data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
    for i in range(len(data['text'])):
        if int(float(data['conf'][i])) > 0:
            # if math.ceil(float(data['conf'][i])) > 50:
            if data['text'][i] == text_data:
                x = data['left'][i]
                y = data['top'][i]
                w = data['width'][i]
                h = data['height'][i]
                return text_data, x, y, w, h


# 获取窗口尺寸
def get_window_size():
    size = pyautogui.size()
    return size


# 截图
def get_picture(image_path, x, y, width, height):
    # img = pyautogui.screenshot(region=[1860, 50, 685, 1200])  # x,y,w,h
    img = pyautogui.screenshot(region=[x, y, width, height])
    # img = Image.fromarray(np.uint8(img))
    # img.save('screenshot3.png')
    img = cv2.cvtColor(np.asarray(img), cv2.COLOR_RGB2BGR)  # cvtColor用于在图像中不同的色彩空间进行转换,用于后续处理。
    cv2.imwrite(image_path, img)


# 删除图片文件
def delete_file(image_path):
    os.remove(image_path)
    print(f'Deleted {image_path}')


# 根据文字获取坐标
def get_pictures_data(text_data):
    while True:
        width = get_window_size()[0]
        height = get_window_size()[1]
        image_path = 'D:\\group_send\\picture_manage\\now_picture.png'
        get_picture(image_path, 0, 0, width, height)
        data = deal_picture(image_path, text_data)
        if data is not None:
            return data


# 唤起企业微信
def wake_up_work_chat():
    ww_window = auto.WindowControl(Name='企业微信', ClassName='WeWorkWindow')
    ww_window.SetFocus()
    ww_window.Maximize(True)


def wake_up_my_client():
    ww_window = auto.WindowControl(Name='use your tool', ClassName='use_your_tool')
    ww_window.SetFocus()


# 循环文件夹内文件夹
def get_folder(main_path):
    files = os.listdir(main_path)
    folder_paths = []
    for file in files:
        file_path = os.path.join(main_path, file)
        if os.path.isdir(file_path):
            folder_paths.append(file_path)
    # print(folder_paths)
    return folder_paths


# 获取企业微信用户名
def get_excel_user(excel_path, sheet):
    try:
        workbook = xlrd.open_workbook(excel_path)
        worksheet = workbook.sheet_by_name(sheet)
        row_count = worksheet.nrows
        user_names = []
        second_user = []
        if (row_count - 1) % 9 != 0:
            second_range = (row_count - 1) // 9 + 1
        else:
            second_range = (row_count - 1) // 9
        for row_num in range(1, row_count):
            user_name = worksheet.cell_value(row_num, 1)
            if user_name is not None:
                second_user.append(user_name)
            if row_num % 9 == 0 or row_num == row_count - 1:
                if second_user:
                    user_names.append(second_user)
                    second_user = []
        # print(second_range, user_names)
        return second_range, user_names
    except Exception as e:
        print(e)


def new_get_excel_user(excel_path, sheet):
    workbook = xlrd.open_workbook(excel_path)
    worksheet = workbook.sheet_by_name(sheet)
    row_count = worksheet.nrows
    user_names = []
    for row_num in range(1, row_count):
        user_name = worksheet.cell_value(row_num, 1)
        if user_name is not None:
            user_names.append(user_name)
    return user_names


# 计算耗时
def count_time(time_str_1, time_str_2):
    # 定义时间格式
    time_format = "%a %b %d %H:%M:%S %Y"

    # 将时间字符串解析为 datetime 对象
    time_obj_1 = datetime.strptime(time_str_1, time_format)
    time_obj_2 = datetime.strptime(time_str_2, time_format)

    # 计算时间差
    time_difference = time_obj_2 - time_obj_1

    # 输出时间差（以秒为单位）
    # print(f"耗时: {time_difference.total_seconds()} 秒")

    # 如果需要更详细的输出（天、小时、分钟、秒）
    days = time_difference.days
    seconds = time_difference.seconds
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return days, hours, minutes, seconds


# 发送结果给群/用户
def send_result_for_me(user_name, forward_count, success_count, start_time, end_time):
    days, hours, minutes, seconds = count_time(start_time, end_time)
    get_picture('D:\\group_send\\picture_manage\\search_box_photo.png', 100, 20, 200, 50)
    # 搜索输入框查询
    search_x, search_y = get_photo_position('D:\\group_send\\picture_manage\\search_box_photo.png')
    pyautogui.moveTo(search_x, search_y)
    pyautogui.click()
    keyboard.write(user_name)
    keyboard.press_and_release('enter')
    time.sleep(1)
    pyautogui.moveTo(search_x + 70, search_y + 70)
    pyautogui.click()
    time.sleep(1)
    if forward_count - success_count > 0:
        text = f'需转发群数量：{forward_count}\n发送成功：{success_count}\n发送失败：{forward_count - success_count}\n失败原因：当前企微无法查到该用户,请维护正确的群名称\n耗时详细：{hours}:{minutes}:{seconds}'
    else:
        text = f'需转发群数量：{forward_count}\n发送成功：{success_count}\n发送失败：{forward_count - success_count}\n耗时详细：{hours}:{minutes}:{seconds}'
    pyperclip.copy(text)
    time.sleep(0.5)
    keyboard.press_and_release('ctrl+v')
    keyboard.press_and_release('enter')


# 获取客户名称和运单号
def get_customer_and_shipment(group_name):
    deal_path = 'D:\\group_send\\ansu_deal\\安速-客服-数据发送.xlsx'
    read_wb = openpyxl.load_workbook(deal_path)
    read_ws = read_wb['运单对应群名']
    read_row_count = get_max_row(read_ws)
    str_txt = ''
    for i in range(2, read_row_count + 1):
        deal_group_name = read_ws.cell(row=i, column=4).value
        if deal_group_name == group_name:
            customer_name = read_ws.cell(row=i, column=1).value
            shipment_num = read_ws.cell(row=i, column=2).value
            str_txt = str_txt + customer_name + ' ' + shipment_num + '\n'
    return str_txt


# 主流程 def main_use_flow(event):
def main_use_flow():
    try:
        start_time = time.asctime(time.localtime(time.time()))
        wake_up_work_chat()
        get_picture('D:\\group_send\\picture_manage\\search_box_photo.png', 100, 15, 200, 50)
        search_sender('D:\\group_send\\picture_manage\\search_box_photo.png', 'D:\\group_send\\picture_manage\\clear_picture.png')
        time.sleep(2)
        text_input_box_x, text_input_box_y = get_photo_position('D:\\group_send\\picture_manage\\text_input_box.png')
        pyautogui.moveTo(text_input_box_x, text_input_box_y+35)
        time.sleep(1)
        pyautogui.click()
        # 发送文本信息
        send_text('D:\\group_send\\upload_file\\消息内容.txt')
        start_text = 'PillowPillowsgxiyPillowPillow'
        end_text = 'PillowPillowakvszPillowPillow'
        files = get_file_name('D:\\group_send\\upload_file')
        send_point(start_text)
        # 发送文件信息
        for file in files:
            send_file(file)
        time.sleep(0.5)
        keyboard.press_and_release('enter')
        send_point(end_text)
        time.sleep(0.5)
        start_point = get_pictures_data(start_text)
        end_point = get_pictures_data(end_text)
        # 输入转发群名称
        deal_excel = 'D:\\group_send\\ansu_deal\\安速-客服-数据发送.xlsx'
        user_info = get_excel_user(deal_excel, "Sheet3")
        for first_index in range(0, user_info[0]):
            # 获取转发内容
            if start_point is not None and end_point is not None:
                # 下拉多选操作
                pyautogui.moveTo(start_point[1] + 10, start_point[2])
                pyautogui.dragTo(end_point[1] + 10, end_point[2], 1, button='left')
                pyautogui.rightClick(start_point[1] + 10, start_point[2])
                pyautogui.moveTo(start_point[1] + 30, start_point[2] + 55)
                pyautogui.click()

                # 选中文本内容 取消选中标识坐标
                time.sleep(1)
                pyautogui.moveTo(start_point[1], start_point[2] - 30)
                pyautogui.click()
                pyautogui.moveTo(start_point[1], start_point[2] + 9)
                pyautogui.click()
                if files == [] or files is None:
                    pyautogui.moveTo(end_point[1], end_point[2] + 20)
                    pyautogui.click()
                time.sleep(1)

                forward_item('D:\\group_send\\picture_manage\\forward_item_pircture.png')
                time.sleep(1)
            if first_index == user_info[0]:
                break
            # 循环用户进行发送
            for second_index in range(0, len(user_info[1][first_index])):
                pyperclip.copy(user_info[1][first_index][second_index])
                time.sleep(0.2)
                keyboard.press_and_release('ctrl+v')
                time.sleep(0.15)
                keyboard.press_and_release('enter')
                time.sleep(0.15)
                # 判断是否能查到用户
                result = get_photo_position('D:\\group_send\\picture_manage\\select_not_exist.png')
                if result is not None:
                    wb = openpyxl.load_workbook(deal_excel)
                    ws = wb['Sheet3']
                    ws.cell(row=second_index + 2 + (first_index) * 9, column=3, value='当前企微无法查到该用户')
                    wb.save(deal_excel)
                    time.sleep(0.2)
                keyboard.press_and_release('ctrl+a')
                time.sleep(0.3)
                second_index += 1
            send_button('D:\\group_send\\picture_manage\\send_button_picture2.png')
        end_time = time.asctime(time.localtime(time.time()))
        # 发送结果
        wb = openpyxl.load_workbook(deal_excel)
        ws = wb['Sheet3']
        forward_count = get_max_row(ws)
        success_count = 0
        for row in range(2, forward_count+1):
            if ws.cell(row, 3).value == '' or ws.cell(row, 3).value is None:
                success_count += 1
                ws.cell(row=row, column=3, value='已发送')
        wb.save(deal_excel)
        send_result_for_me('安速客服-影刀业务对接', forward_count-1, success_count, start_time, end_time)
        time.sleep(0.5)
    except Exception as e:
        print(e)
    finally:
        wake_up_my_client()


def new_main_use_flow():
    try:
        start_time = time.asctime(time.localtime(time.time()))
        wake_up_work_chat()
        deal_excel = 'D:\\group_send\\ansu_deal\\安速-客服-数据发送.xlsx'
        user_names = new_get_excel_user(deal_excel, 'Sheet3')
        for i in range(0, len(user_names)):
            get_picture('D:\\group_send\\picture_manage\\search_box_photo.png', 100, 15, 200, 50)
            user_name = user_names[i]
            print(user_name)
            new_search_sender(user_name, 'D:\\group_send\\picture_manage\\search_box_photo.png')
            time.sleep(2)
            text_input_box_x, text_input_box_y = get_photo_position('D:\\group_send\\picture_manage\\text_input_box.png')
            pyautogui.moveTo(text_input_box_x, text_input_box_y + 35)
            # 发送文本信息
            new_send_text('D:\\group_send\\upload_file\\消息内容.txt', user_name)
            files = get_file_name('D:\\group_send\\upload_file')
            # 发送文件信息
            for file in files:
                send_file(file)
            time.sleep(0.5)
            keyboard.press_and_release('enter')
            time.sleep(0.5)
        end_time = time.asctime(time.localtime(time.time()))
        # 发送结果
        wb = openpyxl.load_workbook(deal_excel)
        ws = wb['Sheet3']
        forward_count = get_max_row(ws)
        success_count = 0
        for row in range(2, forward_count + 1):
            if ws.cell(row, 3).value == '' or ws.cell(row, 3).value is None:
                success_count += 1
                ws.cell(row=row, column=3, value='已发送')
        wb.save(deal_excel)
        send_result_for_me('安速客服-影刀业务对接', forward_count - 1, success_count, start_time, end_time)
    except Exception as e:
        print(e)


def cs_forward_main(base_url, user, pwd):
    # 设置 Tesseract 的路径(仅在必要时，视你的安装情况而定)
    pytesseract.pytesseract.tesseract_cmd = r'D:\group_send\Tesseract-OCR\tesseract.exe'
    # pyautogui禁用故障保护
    pyautogui.FAILSAFE = False
    # pyautogui增加暂停时间
    pyautogui.PAUSE = 0.2
    # 调用键盘
    subprocess.Popen("DrMain", shell=True)
    time.sleep(1)
    if os.path.exists(r'D:\\group_send\\ansu_deal\\安速-客服-数据发送.xlsx'):
        os.remove(r'D:\\group_send\\ansu_deal\\安速-客服-数据发送.xlsx')
    system_download(base_url, user, pwd)
    sales_match_group()
    time.sleep(1)
    new_main_use_flow()


# 客户端 工具页
def tool_ansu_simple_cs_window(tool):
    tool_use_page = tk.Frame(tool, height=884, width=1000, background='white', highlightcolor='black', relief='ridge')
    # 右边功能窗口
    r1 = tk.Label(tool_use_page, text='安速客服 根据运单号 逐个发送', background='white', justify='left')
    r1.place(x=0, y=0)

    d1 = tk.Label(tool_use_page, text='base_url', background='white', justify='left')
    d1.place(x=0, y=25)

    # 设置默认安速系统环境
    e1 = tk.Entry(tool_use_page, width=30, relief='groove')
    e1.insert(0, 'http://aaa-test.fthj-dev.com/')
    e1.place(x=100, y=28)

    d2 = tk.Label(tool_use_page, text='user', background='white')
    d2.place(x=0, y=45)

    e2 = tk.Entry(tool_use_page, width=30, relief='groove')
    e2.place(x=100, y=48)

    d3 = tk.Label(tool_use_page, text='pwd', background='white')
    d3.place(x=0, y=65)

    e3 = tk.Entry(tool_use_page, width=30, relief='groove')
    e3.place(x=100, y=68)

    def clear():
        e1.delete(0, 'end')
        e2.delete(0, 'end')
        e3.delete(0, 'end')

    implement_button = tk.Button(tool_use_page,
                                 text='执行',
                                 background='#AFEEEE',
                                 command=lambda: cs_forward_main(e1.get(), e2.get(), e3.get())
                                 )
    implement_button.place(x=2, y=106)

    clear_button = tk.Button(tool_use_page,
                             text='清空',
                             background='#AFEEEE',
                             command=clear)
    clear_button.place(x=40, y=106)

    # 模块位置
    tool_use_page.place(x=220, y=20)
    return tool_use_page
