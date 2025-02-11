import tkinter as tk
import re
import time
import openpyxl
from datetime import datetime
import keyboard
import uiautomation as auto
import pyperclip
import xlrd
from PIL import Image
import pytesseract, os, pyautogui, cv2, math, random, string, subprocess
import numpy as np
import subprocess

def get_photo_position(image_path):
    try:
        image = pyautogui.locateOnScreen(image_path, grayscale=True)
        center_x, center_y = pyautogui.center(image)
        return center_x, center_y
    except Exception as e:
        print(e, image_path)


# 复制文本信息
def copy_text_file_to_clipboard(file_path):
    with open(file_path, "r", encoding='utf8') as file:
        text = file.read()
        if text is not None:
            pyperclip.copy(text)


# 复制文件
def copy_file_to_clipboard(file_path):
    with open(file_path, "rb") as file:
        image = file.read()
        pyperclip.copy(image)


# 查询发送人信息
def search_sender(search_path, clear_path):
    user_name = '方天研发部-影刀(影刀)'
    # 搜索输入框查询
    search_x, search_y = get_photo_position(search_path)
    pyautogui.moveTo(search_x, search_y)
    pyautogui.click()
    keyboard.write(user_name)
    # keyboard.press_and_release('enter')
    time.sleep(1)
    pyautogui.moveTo(search_x + 70, search_y + 70)
    pyautogui.click()
    time.sleep(0.5)
    # 清空消息记录
    pyautogui.rightClick(search_x + 60, search_y + 60)
    clear_x, clear_y = get_photo_position(clear_path)
    pyautogui.moveTo(clear_x, clear_y)
    pyautogui.click()
    second_confirm_x, second_confirm_y = get_photo_position('D:\\group_send\\picture_manage\\second_confirm.png')
    pyautogui.moveTo(second_confirm_x, second_confirm_y)
    pyautogui.click()


# 点击逐条转发按钮
def forward_item(file_path):
    forward_item_x, forward_item_y = get_photo_position(file_path)
    pyautogui.moveTo(forward_item_x, forward_item_y)
    pyautogui.click()


def position_and_move(file_path):
    send_button_x, send_button_y = get_photo_position(file_path)
    pyautogui.moveTo(send_button_x, send_button_y)
    pyautogui.click()


# 发送按钮
def send_button(file_path):
    send_button_x, send_button_y = get_photo_position(file_path)
    # pyautogui.moveTo(send_button_x, send_button_y)
    pyautogui.moveTo(send_button_x - 150, send_button_y)
    pyautogui.click()


# 发送文本信息
def send_text(file_path):
    copy_text_file_to_clipboard(file_path)
    time.sleep(0.5)
    keyboard.press_and_release('ctrl+v')
    keyboard.press_and_release('enter')


# 循环文件夹内文件 消息内容.txt 除外
def get_file_name(folder_path):
    files = os.listdir(folder_path)
    file_paths = []
    for file in files:
        if re.search('消息内容', file):
            continue
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            file_paths.append(folder_path + "\\" + file)
    # print(file_paths)
    return file_paths


# 发送文件信息
def send_file(file_path):
    if file_path is not None:
        # copy_file_to_clipboard(file_path)
        args = ['powershell', 'Get-Item {} | Set-Clipboard'.format(file_path)]
        subprocess.Popen(args=args, encoding='utf-8')
        time.sleep(0.5)
        keyboard.press_and_release('ctrl+v')
        time.sleep(0.5)


# 输入信息
def send_point(text):
    keyboard.write(text)
    keyboard.press_and_release('enter')


def generate_random_letters(length=5):
    # 选择所有可能的字母（包括大写和小写字母）
    letters = string.ascii_letters.lower()  # 包含 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # 使用 random.choices 随机选择指定数量的字母
    random_letters = ''.join(random.choices(letters, k=length))
    new_letters = 'PillowPillow' + random_letters + 'PillowPillow'
    return new_letters


# 处理图片
def deal_picture(image_path, text_data):
    image = Image.open(image_path)
    pytesseract.load_freq_dawg = False
    pytesseract.load_system_dawg = False
    data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
    for i in range(len(data['text'])):
        if int(float(data['conf'][i])) > 0:
            # if math.ceil(float(data['conf'][i])) > 50:
            if data['text'][i] == text_data:
                x = data['left'][i]
                y = data['top'][i]
                w = data['width'][i]
                h = data['height'][i]
                return text_data, x, y, w, h


# 获取窗口尺寸
def get_window_size():
    size = pyautogui.size()
    return size


# 截图
def get_picture(image_path, x, y, width, height):
    # img = pyautogui.screenshot(region=[1860, 50, 685, 1200])  # x,y,w,h
    img = pyautogui.screenshot(region=[x, y, width, height])
    # img = Image.fromarray(np.uint8(img))
    # img.save('screenshot3.png')
    img = cv2.cvtColor(np.asarray(img), cv2.COLOR_RGB2BGR)  # cvtColor用于在图像中不同的色彩空间进行转换,用于后续处理。
    cv2.imwrite(image_path, img)


# 删除图片文件
def delete_file(image_path):
    os.remove(image_path)
    print(f'Deleted {image_path}')


# 根据文字获取坐标
def get_pictures_data(text_data):
    while True:
        width = get_window_size()[0]
        height = get_window_size()[1]
        image_path = 'D:\\group_send\\picture_manage\\now_picture.png'
        get_picture(image_path, 0, 0, width, height)
        data = deal_picture(image_path, text_data)
        if data is not None:
            return data


# 唤起企业微信
def wake_up_work_chat():
    ww_window = auto.WindowControl(Name='企业微信', ClassName='WeWorkWindow')
    ww_window.SetFocus()
    ww_window.Maximize(True)


def wake_up_my_client():
    ww_window = auto.WindowControl(Name='use your tool', ClassName='use_your_tool')
    ww_window.SetFocus()


# 循环文件夹内文件夹
def get_folder(main_path):
    files = os.listdir(main_path)
    folder_paths = []
    for file in files:
        file_path = os.path.join(main_path, file)
        if os.path.isdir(file_path):
            folder_paths.append(file_path)
    # print(folder_paths)
    return folder_paths


# 获取企业微信用户名
def get_excel_user(excel_path, sheet):
    try:
        workbook = xlrd.open_workbook(excel_path)
        worksheet = workbook.sheet_by_name(sheet)
        row_count = worksheet.nrows
        user_names = []
        second_user = []
        if (row_count - 1) % 9 != 0:
            second_range = (row_count - 1) // 9 + 1
        else:
            second_range = (row_count - 1) // 9
        for row_num in range(1, row_count):
            user_name = worksheet.cell_value(row_num, 1)
            if user_name is not None:
                second_user.append(user_name)
            if row_num % 9 == 0 or row_num == row_count - 1:
                if second_user:
                    user_names.append(second_user)
                    second_user = []
        # print(second_range, user_names)
        return second_range, user_names
    except Exception as e:
        print(e)


# 计算耗时
def count_time(time_str_1, time_str_2):
    # 定义时间格式
    time_format = "%a %b %d %H:%M:%S %Y"

    # 将时间字符串解析为 datetime 对象
    time_obj_1 = datetime.strptime(time_str_1, time_format)
    time_obj_2 = datetime.strptime(time_str_2, time_format)

    # 计算时间差
    time_difference = time_obj_2 - time_obj_1

    # 输出时间差（以秒为单位）
    # print(f"耗时: {time_difference.total_seconds()} 秒")

    # 如果需要更详细的输出（天、小时、分钟、秒）
    days = time_difference.days
    seconds = time_difference.seconds
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return days, hours, minutes, seconds


# 发送结果给群/用户
def send_result_for_me(user_name, forward_count, success_count, start_time, end_time):
    days, hours, minutes, seconds = count_time(start_time, end_time)
    get_picture('D:\\group_send\\picture_manage\\search_box_photo.png', 100, 20, 200, 50)
    # 搜索输入框查询
    search_x, search_y = get_photo_position('D:\\group_send\\picture_manage\\search_box_photo.png')
    pyautogui.moveTo(search_x, search_y)
    pyautogui.click()
    keyboard.write(user_name)
    keyboard.press_and_release('enter')
    time.sleep(1)
    pyautogui.moveTo(search_x + 70, search_y + 70)
    pyautogui.click()
    time.sleep(1)
    if forward_count - success_count > 0:
        text = f'需转发群数量：{forward_count}\n发送成功：{success_count}\n发送失败：{forward_count - success_count}\n失败原因：当前企微无法查到该用户,请维护正确的群名称\n耗时详细：{hours}:{minutes}:{seconds}'
    else:
        text = f'需转发群数量：{forward_count}\n发送成功：{success_count}\n发送失败：{forward_count - success_count}\n耗时详细：{hours}:{minutes}:{seconds}'
    pyperclip.copy(text)
    time.sleep(0.5)
    keyboard.press_and_release('ctrl+v')
    keyboard.press_and_release('enter')


# 主流程 def main_use_flow(event):
def main_use_flow():
    try:
        start_time = time.asctime(time.localtime(time.time()))
        wake_up_work_chat()
        get_picture('D:\\group_send\\picture_manage\\search_box_photo.png', 100, 20, 200, 50)
        search_sender('D:\\group_send\\picture_manage\\search_box_photo.png', 'D:\\group_send\\picture_manage\\clear_picture.png')
        time.sleep(2)
        text_input_box_x, text_input_box_y = get_photo_position('D:\\group_send\\picture_manage\\text_input_box.png')
        pyautogui.moveTo(text_input_box_x, text_input_box_y+35)
        time.sleep(1)
        pyautogui.click()
        # 发送文本信息
        send_text('D:\\group_send\\upload_file\\消息内容.txt')
        start_text = 'PillowPillowsgxiyPillowPillow'
        end_text = 'PillowPillowakvszPillowPillow'
        files = get_file_name('D:\\group_send\\upload_file')
        send_point(start_text)
        # 发送文件信息
        for file in files:
            send_file(file)
        time.sleep(0.5)
        keyboard.press_and_release('enter')
        send_point(end_text)
        time.sleep(0.5)
        start_point = get_pictures_data(start_text)
        end_point = get_pictures_data(end_text)
        # 输入转发群名称
        deal_excel = 'D:\\group_send\\forward_users.xlsx'
        # 设置发送结果为空行
        wb = openpyxl.load_workbook(deal_excel)
        ws = wb['Sheet1']
        ws.delete_cols(3)
        ws.cell(row=1, column=3, value='发送结果')
        wb.save(deal_excel)
        user_info = get_excel_user(deal_excel, "Sheet1")
        for first_index in range(0, user_info[0]):
            # 获取转发内容
            if start_point is not None and end_point is not None:
                # 下拉多选操作
                pyautogui.moveTo(start_point[1] + 10, start_point[2])
                pyautogui.dragTo(end_point[1] + 10, end_point[2], 1, button='left')
                pyautogui.rightClick(start_point[1] + 10, start_point[2])
                pyautogui.moveTo(start_point[1] + 30, start_point[2] + 55)
                pyautogui.click()

                # 选中文本内容 取消选中标识坐标
                time.sleep(1)
                pyautogui.moveTo(start_point[1], start_point[2] - 30)
                pyautogui.click()
                pyautogui.moveTo(start_point[1], start_point[2] + 9)
                pyautogui.click()
                if files == [] or files is None:
                    pyautogui.moveTo(end_point[1], end_point[2] + 20)
                    pyautogui.click()
                time.sleep(1)

                forward_item('D:\\group_send\\picture_manage\\forward_item_pircture.png')
                time.sleep(1)
            if first_index == user_info[0]:
                break
            # 循环用户进行发送
            for second_index in range(0, len(user_info[1][first_index])):
                pyperclip.copy(user_info[1][first_index][second_index])
                time.sleep(0.2)
                keyboard.press_and_release('ctrl+v')
                time.sleep(0.15)
                keyboard.press_and_release('enter')
                time.sleep(0.15)
                # 判断是否能查到用户
                result = get_photo_position('D:\\group_send\\picture_manage\\select_not_exist.png')
                if result is not None:
                    wb = openpyxl.load_workbook(deal_excel)
                    ws = wb['Sheet1']
                    ws.cell(row=second_index+2+(first_index)*9, column=3, value='当前企微无法查到该用户')
                    wb.save(deal_excel)
                    time.sleep(0.2)
                keyboard.press_and_release('ctrl+a')
                time.sleep(0.3)
                second_index += 1
            send_button('D:\\group_send\\picture_manage\\send_button_picture2.png')
        end_time = time.asctime(time.localtime(time.time()))


        # 发送结果
        wb = openpyxl.load_workbook(deal_excel)
        ws = wb['Sheet1']
        i = ws.max_row
        forward_count = 0
        while ws.max_row > 0:
            row_dict = {i.value for i in ws[i]}
            if row_dict == {None}:
                i = i - 1
            else:
                forward_count = i
                break
        success_count = 0
        for row in range(2, forward_count+1):
            if ws.cell(row, 3).value == '' or ws.cell(row, 3).value is None:
                success_count += 1
                ws.cell(row=row, column=3, value='已发送')
        wb.save(deal_excel)
        send_result_for_me('影刀项目组', forward_count-1, success_count, start_time, end_time)
        time.sleep(0.5)
    except Exception as e:
        print(e)
    finally:
        wake_up_my_client()


def forward_everyone_main():
    # 设置 Tesseract 的路径(仅在必要时，视你的安装情况而定)
    # pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    pytesseract.pytesseract.tesseract_cmd = r'D:\group_send\Tesseract-OCR\tesseract.exe'
    # pyautogui禁用故障保护
    pyautogui.FAILSAFE = False
    # pyautogui增加暂停时间
    pyautogui.PAUSE = 0.2
    # 调用键盘
    subprocess.Popen("DrMain", shell=True)
    time.sleep(1.5)

    main_use_flow()


# 客户端 工具页
def tool_forward_everyone_window(tool):
    tool_use_page = tk.Frame(tool, height=884, width=1000, background='white', highlightcolor='black', relief='ridge')
    # 右边功能窗口
    r1 = tk.Label(tool_use_page, text='企业微信群转发', background='white', justify='left')
    r1.place(x=0, y=0)

    # 说明内容
    describe = ('说明：该功能用于本地微信群发，执行前须知：\n'
                '1、电脑登录企业微信；\n'
                '2、在对应路径维护发送信息；\n'
                '3、在对应xlsx维护发送人；\n'
                '4、执行中不要操作键盘和鼠标避免转发内容出错')

    d3 = tk.Label(tool_use_page, text=describe, background='white', justify='left')
    d3.place(x=0, y=25)

    # 执行按钮
    implement_button = tk.Button(tool_use_page,
                                 text='执行',
                                 background='#AFEEEE',
                                 command=forward_everyone_main)
    implement_button.place(x=2, y=126)

    tool_use_page.place(x=220, y=20)
    return tool_use_page




